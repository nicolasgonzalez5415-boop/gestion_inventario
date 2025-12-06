import csv
import os
import numpy as np
from datetime import datetime, timedelta
import openpyxl
import matplotlib.pyplot as plt
from openpyxl.utils import get_column_letter
import streamlit as st
import pandas as pd
import gspread

st.set_page_config(layout="wide", page_title="Inventario B&M")
st.title("Sistema de Gestión de Inventario B&M")
inventario_xls = 'inventario.xlsx'
stock_minimo_xls  = 'stock_minimo.xlsx'
movimientos_xls = 'movimientos.xlsx'

inventario = {}
stock_minimo = {}
movimientos = []
movimientos_headers = ["timestamp", "tipo", "codigo", "nombre", "cantidad", "fecha_vencimiento", "precio_costo", "precio_venta"]
stock_minimo_headers = ['codigo', 'stock_min']
inventario_headers = ["codigo", "nombre", "marca", "cantidad", "fecha_vencimiento", "precio_costo", "precio_venta"]

# --- CONFIGURACIÓN DE LA PÁGINA ---
st.set_page_config(layout="wide", page_title="Inventario B&M")
st.title("Sistema de Gestión de Inventario B&M (Google Sheets)")

# --- CONFIGURACIÓN GOOGLE SHEETS ---
# USAR ID EN LUGAR DE NOMBRE (Más seguro)
# Extraído de tu link: docs.google.com/spreadsheets/d/1Zu-Dq6UCYRKMTWNsxj8FsMzzpAdtvl-qb40CVEmwl44/...
GOOGLE_SHEET_ID = "1Zu-Dq6UCYRKMTWNsxj8FsMzzpAdtvl-qb40CVEmwl44"

# Nombres de las pestañas
INVENTARIO_WS = 'inventario'
STOCK_MINIMO_WS = 'stock_minimo'
MOVIMIENTOS_WS = 'movimientos'

@st.cache_resource(ttl=3600)
def obtener_conexion():
    """Conecta con Google Sheets usando st.secrets"""
    try:
        # Crea un diccionario con las credenciales desde secrets
        credentials = dict(st.secrets["gcp_service_account"])
        
        # PARCHE CRÍTICO: Reemplazar saltos de línea literales por reales
        # Esto soluciona errores de autenticación comunes en Streamlit
        if "private_key" in credentials:
            credentials["private_key"] = credentials["private_key"].replace("\\n", "\n")
        
        # Autentica
        gc = gspread.service_account_from_dict(credentials)
        
        # Abre la hoja POR ID (Mucho más robusto que por título)
        sh = gc.open_by_key(GOOGLE_SHEET_ID)
        return sh
    except Exception as e:
        st.error(f"Error de conexión: {e}. \n\nPosibles causas:\n1. El bot no tiene permiso de 'Editor' en la hoja.\n2. La API de Google Sheets no está habilitada.")
        st.stop()

def iniciarlizar_archivos():
    try:
        if not os.path.exists(inventario_xls):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'datos'
            ws.append(inventario_headers)
            wb.save(inventario_headers)
            wb.close()

        if not os.path.exists(movimientos_xls):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'datos'
            ws.append(movimientos_headers)
            wb.save(movimientos_xls)
            wb.close()
    except Exception as e:
        st.error(f"Error: {e}")


def normalizar_fecha(fecha_obj) -> str:
    if not fecha_obj:
        return ""
    try:
        if isinstance(fecha_obj, str):
            fecha_obj = fecha_obj.strip()
            if ' ' in fecha_obj:
                fecha_obj = fecha_obj.split(' ')[0]
            if 'T' in fecha_obj:
                fecha_obj = fecha_obj.split('T')[0]
            return fecha_obj
        
        if hasattr(fecha_obj, 'strftime'):
            return fecha_obj.strftime("%Y-%m-%d")
        
        return str(fecha_obj).split(' ')[0]
    except:
        return ""

def _convertir_a_numero(valor, por_defecto=0):
    if valor is None or valor == '':
        return por_defecto
    try:
        return int(valor)
    except (ValueError, TypeError):
        try:
            return float(valor)
        except (ValueError, TypeError):
            return por_defecto
        

def stock_total(codigo: str) -> int:
    return sum(l["cantidad"] for l in inventario.get(codigo, []))

def ordenar_lotes_fifo(lotes):
    def clave(l):
        fv = l.get("fecha_vencimiento", "")
        if not fv:
            return (datetime.max, 1)
        try:
            return (datetime.fromisoformat(fv), 0)
        except Exception:
            return (datetime.max, 1)
    return sorted(lotes, key=clave)

def _escribir_xlsx(ruta, headers, filas):
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'datos'
        
        ws.append(headers)

        for fila in filas:
            fila_limpia = []
            for celda in fila:
                if isinstance(celda, (datetime, pd.Timestamp)):
                    fila_limpia.append(celda.isoformat())
                else:
                    fila_limpia.append(celda)
            ws.append(fila_limpia)
        
        for c_idx, column_cells in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(c_idx)
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max(12, min(50, max_length + 2))
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(ruta)
        wb.close()
    except PermissionError:
         print(f"Error al guardar {ruta}.")
    except Exception as e:
        print(f"Error al guardar {ruta}: {e}")

def cargar_stock_min():
    stock_minimo.clear()
    if not os.path.exists(stock_minimo_xls):
        return
    
    try:
        wb = openpyxl.load_workbook(stock_minimo_xls, data_only = True)
        ws = wb.active
    except Exception:
        return

    for fila in ws.iter_rows(min_row=2, values_only=True):
        if not fila or fila[0] is None:
            continue
        
        try:
            codigo = str(fila[0]).strip()
            minimo = _convertir_a_numero(fila[1], por_defecto=0)
            if codigo:
                stock_minimo[codigo] = minimo
        except Exception:
            return
    wb.close()

def cargar_inventario():
    inventario.clear()
    if not os.path.exists(inventario_xls):
        return
    
    try:
        wb = openpyxl.load_workbook(inventario_xls, data_only = True)
        ws = wb.active
    except Exception:
        return

    for fila in ws.iter_rows(min_row=2, values_only=True):
        if not fila or fila[0] is None:
            continue
        
        try:
            datos = list(fila) + [None]*(7 - len(fila))
            codigo, nombre, marca, cantidad, fecha_vencimiento, precio_costo, precio_venta = datos[:7]
        except ValueError as e:
            continue

        cantidad = _convertir_a_numero(cantidad, por_defecto=0)
        precio_costo = _convertir_a_numero(precio_costo, por_defecto=0)
        precio_venta = _convertir_a_numero(precio_venta, por_defecto=0)
        
        fecha_str = normalizar_fecha(fecha_vencimiento)
        
        lote = {
            'nombre': str(nombre),
            'marca': str(marca),
            'cantidad': cantidad,
            'fecha_vencimiento': fecha_str,
            'precio_costo': precio_costo,
            'precio_venta': precio_venta
        }

        codigo_str = str(codigo).strip()
        if codigo_str not in inventario:
            inventario[codigo_str] = []
        inventario[codigo_str].append(lote)
    wb.close()

def guardar_inventario():
    
    filas = []
    for codigo, lotes in inventario.items():
        for d in lotes:
            filas.append([
                codigo,
                d.get('nombre',""),
                d.get('marca',""),
                d.get('cantidad',0),
                d.get('fecha_vencimiento',""),
                d.get('precio_costo',""),
                d.get('precio_venta',""),
            ])
    _escribir_xlsx(inventario_xls, inventario_headers, filas)

def guardar_stock_minimo():
    filas = []
    for codigo, minimo in stock_minimo.items():
        filas.append([codigo, minimo])

    _escribir_xlsx(stock_minimo_xls, stock_minimo_headers, filas)


def cargar_movimientos():
    movimientos.clear()
    if not os.path.exists(movimientos_xls):
        return

    try:
        wb = openpyxl.load_workbook(movimientos_xls, data_only = True)
        ws = wb.active
    except Exception:
        return
    
    for fila_tupla in ws.iter_rows(min_row=2, values_only=True):
        if not fila_tupla or not fila_tupla[0]:
            continue
        movimientos.append(list(fila_tupla))
        
    wb.close()    
        
def guardar_movimientos():
    _escribir_xlsx(movimientos_xls, movimientos_headers, movimientos)


def registrar_movimiento(tipo, codigo, nombre, cantidad, fecha_vencimiento, precio_costo, precio_venta):
    nueva_fila = [
        datetime.now().isoformat(timespec="seconds"),
        tipo,
        str(codigo),
        nombre,
        cantidad,
        fecha_vencimiento or "",
        precio_costo if precio_costo is not None else 0,
        precio_venta if precio_venta is not None else 0,
    ]
    movimientos.append(nueva_fila)
    guardar_movimientos()


iniciarlizar_archivos()
cargar_inventario()
cargar_movimientos()
cargar_stock_min()


tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs(["Registrar Entrada", "Registrar Salida", "Mostrar Inventario", "Reporte de Movimientos", "Reporte de Niveles de Stock", "Reporte de Alertas de Vencimiento"])

with tab1:
    st.subheader("Registro de Entradas: ")
    
    if 'reset_counter' not in st.session_state:
        st.session_state.reset_counter = 0

    input_key = f"entrada_input_{st.session_state.reset_counter}"

    st.markdown(f"""
        <script>
        function focusInput() {{
            const inputs = window.parent.document.querySelectorAll('input[type="text"]');
            if (inputs.length > 0) {{ inputs[0].focus(); }}
        }}
        setTimeout(focusInput, 200);
        </script>
    """, unsafe_allow_html=True)


    entrada = st.text_input("Escanee el código o esciba 'buscar': ", key=input_key)
    
    codigo_seleccionado = None
    
    if entrada:
        if entrada.lower() == 'buscar':
            productos_lista = []
            for codigo, lotes in inventario.items():
                if lotes:
                    base = lotes[0]
                    nombre = base.get('nombre') or 'N/A'
                    marca = base.get('marca') or 'N/A'
                    productos_lista.append((codigo, nombre, marca))
                
            productos_lista.sort(key=lambda x: x[1])

            opciones = [f"{i+1}) {nombre} - {marca} (Código: {codigo})" 
                                for i, (codigo, nombre, marca) in enumerate(productos_lista)]
            
            opciones.insert(0, "Cancelar")
            seleccion = st.selectbox("Seleccione un producto", opciones)

            if seleccion == 'Cancelar':
                st.info("Selección cancelada")
                codigo_seleccionado = None
            else:
                idx = opciones.index(seleccion) - 1
                codigo_seleccionado = productos_lista[idx][0]
                st.success(f"Se seleleccionó: {productos_lista[idx][1]} (Codigo: {codigo})")
        else:
            codigo_seleccionado = entrada

    if codigo_seleccionado:
        es_nuevo = codigo_seleccionado not in inventario
        if codigo_seleccionado in stock_minimo:
            no_tiene_min = stock_minimo[codigo_seleccionado] is None
        else:
            no_tiene_min = True

        if es_nuevo:
            st.info(f"El código {codigo_seleccionado} no se encuentra en el inventario. Se creará un nuevo producto")
        else:
            base = inventario[codigo_seleccionado][0]
            st.success(f"Producto existente: {base.get('nombre')} ({base.get('marca')})")
        
        with st.form("form_entrada", clear_on_submit = True):
            nombre_def = '' if es_nuevo else inventario[codigo_seleccionado][0].get('nombre', '')
            marca_def = '' if es_nuevo else inventario[codigo_seleccionado][0].get('marca', '')
            pc_def = 0 if es_nuevo else inventario[codigo_seleccionado][0].get('precio_costo', 0)
            pv_def = 0 if es_nuevo else inventario[codigo_seleccionado][0].get('precio_venta', 0)
            cant_min_def = 0 if no_tiene_min else stock_minimo[codigo_seleccionado]
            
            nombre = st.text_input("Nombre", value = nombre_def, disabled = not es_nuevo)
            marca = st.text_input("Marca", value = marca_def, disabled = not es_nuevo)
            precio_costo = st.number_input("Precio de Costo", min_value = 0, value = int(pc_def), disabled = not es_nuevo)
            precio_venta = st.number_input("Precio de Venta", min_value = 0, value = int(pv_def), disabled = not es_nuevo)

            cantidad = st.number_input("Cantidad a ingresar", min_value = 0, value = 1, step = 1)
            cant_min = st.number_input("Cantidad mínima", min_value = 0, value = int(cant_min_def))
            aplica_vencimiento = st.checkbox("¿El producto tiene fecha de vencimiento?", value = True)
            fecha_vencimiento = st.date_input("Fecha Vencimiento", value = datetime.now().date())

            submitted = st.form_submit_button("Guardar")

            if submitted:
                fv = normalizar_fecha(fecha_vencimiento) if aplica_vencimiento else ""

                if es_nuevo:
                    lote = {
                        'nombre': nombre,
                        'marca': marca,
                        'cantidad': cantidad,
                        'fecha_vencimiento': fv,
                        'precio_costo': precio_costo,
                        'precio_venta': precio_venta
                    }

                    inventario[codigo_seleccionado] = [lote]
                    stock_minimo[codigo_seleccionado] = cant_min
                    mensaje = f'Producto {nombre} creado con éxito'
                else:
                    lotes = inventario[codigo_seleccionado]
                    lote_existente = next((l for l in lotes if l.get("fecha_vencimiento", "") == fv), None)

                    if lote_existente:
                        lote_existente['cantidad'] += cantidad
                        mensaje = f"Se agregaron {cantidad} unidades al lote existente ({fv})"
                    else:
                        lotes.append({
                            'nombre': nombre,
                            'marca': marca,
                            'cantidad': cantidad,
                            'fecha_vencimiento': fv,
                            'precio_costo': precio_costo,
                            'precio_venta': precio_venta
                        })
                        stock_minimo[codigo_seleccionado] = cant_min
                        mensaje = f"Se creó un nuevo lote con {cantidad} unidades ({fv})"

                guardar_inventario()
                guardar_stock_minimo()
                registrar_movimiento("entrada", codigo_seleccionado, nombre, cantidad, fv, precio_costo, precio_venta)
                st.success(mensaje) 
                st.session_state.reset_counter += 1
                st.rerun()

with tab2:

    if "lista" not in st.session_state:
        st.session_state.lista = {}

    def procesar_codigo_escaneado():
        codigo_sin_procesar = st.session_state.codigo
        if not codigo_sin_procesar:
            return

        cantidad = 1
        codigo_producto = ""

        if "*" in codigo_sin_procesar:
            try:
                partes = codigo_sin_procesar.split("*", 1)
                cantidad = int(partes[0].strip())
                codigo_producto = partes[1].strip()
            except:
                codigo_producto = codigo_sin_procesar.strip()
        else:
            codigo_producto = codigo_sin_procesar.strip()
            cantidad = 1
        
        if codigo_producto not in inventario:
            st.toast(f"El {codigo_producto} no existe")
            st.session_state.codigo = ""
            return

        stock_disp = stock_total(codigo_producto)
        en_lista = st.session_state.lista.get(codigo_producto, 0)
        
        if (en_lista + cantidad) > stock_disp:
            st.toast(f"Stock insuficiente. Disponible: {stock_disp}")
        else:
            if codigo_producto in st.session_state.lista:
                st.session_state.lista[codigo_producto] += cantidad
            else:
                st.session_state.lista[codigo_producto] = cantidad
            st.toast(f"Agregado: {codigo_producto}")

        st.session_state.codigo = ""

    def enfoque_automatico():
        st.markdown(
            """
        <script>
        function focusInput(){
            try {
                const input = window.parent.document.querySelector('input[aria-label="Escanee el código del producto (Si son varios, ej: 5*7806505055391)"]');
                if (input) {
                    input.focus();
                }
            } catch(e) { console.error('Error al enfocar:', e); }
        }
        setTimeout(focusInput, 100);
        </script>
        """,
        unsafe_allow_html=True
        )

    st.subheader("Registro de Salidas")
    st.text_input("Escanee el código del producto (Si son varios, ej: 5*7806505055391)", key="codigo", on_change=procesar_codigo_escaneado)
    enfoque_automatico()

    st.divider()
    st.subheader("Lista Actual de Productos")

    if not st.session_state.lista:
        st.info("No hay productos en la lista")
    else:
        total_productos = 0
        for codigo, cant_lista in st.session_state.lista.items():
            if codigo in inventario:
                nombre = inventario[codigo][0]['nombre']
                st.write(f"- **{nombre}**: {cant_lista} unidades")
                total_productos += cant_lista
            
        st.subheader(f"Total de Productos: {total_productos}")

        if st.button("Registrar Salidas (Confirmar)"):
            for codigo_prod, cantidad_sacar in st.session_state.lista.items():
                if codigo_prod not in inventario: continue

                lotes_a_modificar = ordenar_lotes_fifo(inventario[codigo_prod])
                restante = cantidad_sacar
                nombre_prod = lotes_a_modificar[0]['nombre']

                lotes_finales = []
                for l in lotes_a_modificar:
                    if restante > 0:
                        toma = min(l['cantidad'], restante)
                        l['cantidad'] -= toma
                        restante -= toma
                        
                        registrar_movimiento("salida", codigo_prod, nombre_prod, toma, l.get('fecha_vencimiento',''), l.get('precio_costo'), l.get('precio_venta'))
                    
                    if l['cantidad'] > 0:
                        lotes_finales.append(l)
                
                if not lotes_finales:
                    if codigo_prod in inventario:
                        del inventario[codigo_prod]
                else:
                    inventario[codigo_prod] = lotes_finales
            
            guardar_inventario()
            st.session_state.lista = {}
            st.success("Salidas registradas correctamente")
            st.rerun() 

with tab3:
    st.subheader("Inventario Completo")
    if os.path.exists(inventario_xls):
        try:
            df_inv = pd.read_excel(inventario_xls, engine='openpyxl')
            
            busqueda = st.text_input("🔍 Buscar producto (Nombre, Marca o Código):", key="search_inv")

            if busqueda:
                codigo = df_inv['codigo'].astype(str).str.contains(busqueda, case=False, na=False)
                df_inv = df_inv[codigo]

            
            st.dataframe(
                df_inv, 
                width="stretch", 
                height=500,
                hide_index=True,
                column_config={
                    'codigo': st.column_config.TextColumn("Código"),
                    'nombre': st.column_config.TextColumn("Nombre"),
                    'marca': st.column_config.TextColumn("Marca"),
                    "fecha_vencimiento": st.column_config.TextColumn("Fecha de Vencimiento"),
                    "cantidad": st.column_config.NumberColumn("Stock"),
                    "precio_costo": st.column_config.NumberColumn("P. Costo", format="$%d"),
                    "precio_venta": st.column_config.NumberColumn("P. Venta", format="$%d"),
                }
            )
            
            st.caption(f"Stock Total: {stock_total(busqueda)}")

        except Exception as e:
            st.error(f"Error al leer archivo Excel: {e}")
    else:
        st.warning("No hay archivo de inventario.")

with tab4:
    st.subheader("Reporte de Movimientos: ")
    col_izq, col_der = st.columns(2)

    with col_izq:

        nombre_seleccionado = None
        productos_lista = []
        for codigo, lotes in inventario.items():
            if lotes:
                base = lotes[0]
                nombre = base.get('nombre') or 'N/A'
                marca = base.get('marca') or 'N/A'
                productos_lista.append((codigo, nombre, marca))
            
        productos_lista.sort(key=lambda x: x[1])

        opciones = [f"{i+1}) {nombre} - {marca} (Código: {codigo})" 
                            for i, (codigo, nombre, marca) in enumerate(productos_lista)]
        
        opciones.insert(0, "Cancelar")
        opciones.insert(1, 'Todos los productos')
        seleccion = st.selectbox("Seleccione un producto", opciones)


        if seleccion == 'Cancelar':
            st.info("Selección cancelada")
        elif seleccion == 'Todos los productos':
            nombre_seleccionado = None
        else:
            idx = opciones.index(seleccion) - 2
            codigo = productos_lista[idx][0]
            nombre_seleccionado = productos_lista[idx][1]
            st.success(f"Se seleleccionó: {productos_lista[idx][1]} (Codigo: {codigo})")

        fecha_inicio = st.date_input("Ingrese la fecha de inicio: ")
        fecha_fin = st.date_input("Ingrese la fecha de fin: ")

        tipo_movimiento = st.multiselect("Seleccione el tipo de movimiento", ["Entrada", "Salida"], key="tipo_movimiento")

    if col_izq.button("Mostrar Movimientos: "):
        try:
            df = pd.read_excel(movimientos_xls)
        
            
            df['timestamp'] = pd.to_datetime(df['timestamp'])
            fecha_inicio = pd.to_datetime(fecha_inicio)
            fecha_fin = pd.to_datetime(fecha_fin) + pd.Timedelta(days=1)
            df_filtrado = df.loc[(df["timestamp"] >= fecha_inicio) & (df["timestamp"] < fecha_fin)]
            df_filtrado['datetime'] = df_filtrado['timestamp']

            tipo_movimiento = [tipo.lower() for tipo in tipo_movimiento]

            if tipo_movimiento:
                df_filtrado = df_filtrado[df_filtrado['tipo'].isin(tipo_movimiento)]

            if nombre_seleccionado is not None:
                df_filtrado = df_filtrado[df_filtrado['nombre'] == nombre_seleccionado]

            df_filtrado['fecha'] = df_filtrado['timestamp'].dt.date
            df_filtrado['hora'] = df_filtrado['timestamp'].dt.time


            columnas_finales = ["fecha", 'hora', "tipo", "nombre", "cantidad"]
            columnas_finales = [c for c in columnas_finales if c in df_filtrado.columns]

            df_filtrado = df_filtrado.rename(columns = {'fecha': 'Fecha', 'hora': "Hora", "tipo": "Tipo", "nombre": "Nombre", "cantidad": "Cantidad"})
            
            nuevos_nombres = ["Fecha", 'Hora', "Tipo", "Nombre", "Cantidad"]
            nuevos_nombres = [c for c in nuevos_nombres if c in df_filtrado.columns]


            if 'entrada' in tipo_movimiento and 'salida' in tipo_movimiento:
                col_izq.write("Entrada y Salida")
                df_e = df_filtrado[df_filtrado['Tipo'] == 'entrada']

                fig1, ax1 = plt.subplots(figsize=(10, 5))
                ax1.scatter(df_e["timestamp"], df_e["Cantidad"], alpha=0.7, edgecolor="k")

                ax1.set_xlabel("Fecha / Hora")
                ax1.set_ylabel("Cantidad")
                ax1.set_title(f"Entradas de {productos_lista[idx][1]}")
                ax1.grid(True)
                col_der.pyplot(fig1)
                
                df_s = df_filtrado[df_filtrado['Tipo'] == 'salida']

                fig2, ax2 = plt.subplots(figsize=(10, 5))
                ax2.scatter(df_s["timestamp"], df_s["Cantidad"], alpha=0.7, edgecolor="k")

                ax2.set_xlabel("Fecha / Hora")
                ax2.set_ylabel("Cantidad")
                ax2.set_title(f"Salidas de {productos_lista[idx][1]}")
                ax2.grid(True)
                col_der.pyplot(fig2)

            elif 'salida' in tipo_movimiento:
                col_izq.write("Salida")
                df_s = df_filtrado[df_filtrado['Tipo'] == 'salida']

                fig, ax = plt.subplots(figsize=(10, 5))
                ax.scatter(df_s["timestamp"], df_s["Cantidad"], alpha=0.7, edgecolor="k")

                ax.set_xlabel("Fecha / Hora")
                ax.set_ylabel("Cantidad")
                ax.set_title(f"Salidas de {productos_lista[idx][1]}")
                ax.grid(True)
                col_der.pyplot(fig)
            else:
                col_izq.write("Entrada")
                df_e = df_filtrado[df_filtrado['Tipo'] == 'entrada']

                fig, ax = plt.subplots(figsize=(10, 5))
                ax.scatter(df_e["timestamp"], df_e["Cantidad"], alpha=0.7, edgecolor="k")

                ax.set_xlabel("Fecha / Hora")
                ax.set_ylabel("Cantidad")
                ax.set_title(f"Entradas de {productos_lista[idx][1]}")
                ax.grid(True)
                col_der.pyplot(fig)
        except:
            pass

        if nombre_seleccionado is not None:
            col_izq.markdown("### Tabla de Movimientos")
            col_izq.table(df_filtrado[nuevos_nombres])
        else:
            col_der.markdown("### Tabla de Movimientos")
            col_der.table(df_filtrado[nuevos_nombres])
        

with tab5:
    st.subheader("Reporte de Niveles de Stock: ")
    if os.path.exists(stock_minimo_xls):
        try:
            df_stock_min = pd.read_excel(stock_minimo_xls, engine='openpyxl')
            df_inv = pd.read_excel(inventario_xls, engine='openpyxl')
            
            df_inv_sin_lotes = df_inv.groupby(['codigo', 'nombre'], as_index=False)['cantidad'].sum()
            df_inv_sin_lotes = df_inv_sin_lotes.rename(columns = {'cantidad': 'stock_total_calc'})

            df_reporte = pd.merge(
                df_stock_min,
                df_inv_sin_lotes[['codigo', 'nombre', 'stock_total_calc']],
                on = 'codigo',
                how = 'left'
            )

            
            df_reporte['stock_total_calc'] = df_reporte['stock_total_calc'].fillna(0)
            df_reporte['nombre'] = df_reporte['nombre'].fillna('Sin nombre')


            def semaforo(fila):
                total = fila['stock_total_calc']
                minimo = fila['stock_min']

                if total <= minimo:
                    return "🔴"
                elif total <= 1.5*minimo:
                    return "🟡"
                else:
                    return "🟢"

            df_reporte['semaforo'] = df_reporte.apply(semaforo, axis = 1)

            busqueda = st.text_input("🔍 Buscar producto (Código):", key="search_stock_min")

            if busqueda:
                codigo = df_reporte['codigo'].astype(str).str.contains(busqueda, case=False, na=False)
                df_reporte = df_reporte[codigo]
            

            st.dataframe(
                df_reporte, 
                width="stretch", 
                height="stretch",
                hide_index=True,
                column_config={
                    'codigo': st.column_config.TextColumn("Código"),
                    'nombre': st.column_config.TextColumn("Nombre"),
                    "stock_min": st.column_config.NumberColumn("Stock Mínimo"),
                    'stock_total_calc': st.column_config.NumberColumn("Stock Total Calculado"),
                    'semaforo': st.column_config.TextColumn("Estado")
                },
                column_order=("codigo", 'nombre', 'stock_min', 'stock_total_calc', 'semaforo')
            )
            
        except Exception as e:
            st.error(f"Error al leer archivo Excel: {e}")
    else:
        st.warning("No hay archivo de inventario.")


with tab6:
    st.subheader("Reporte de Alertas de Vencimiento: ")
    
    alerta_critica = st.slider("Ingrese los días para las Alertas Críticas", 0, 130, 3)
    alerta_adv = st.slider("Ingrese los días para las Alertas de Advertencia", 0, 130, 7)
    alerta_preventiva = st.slider("Ingrese los días para las Alertas Preventivas", 0, 130, 12)

    hoy = datetime.now().date()

    limite_crit = hoy + timedelta(days=alerta_critica)
    limite_adv = hoy + timedelta(days=alerta_adv)
    limite_prev = hoy + timedelta(days=alerta_preventiva)

    alertas = []

    for codigo, lotes in inventario.items():
        for lote in lotes:
            fv = lote.get("fecha_vencimiento")
            estado = None
            if fv:
                try:
                    fv_date = datetime.strptime(fv, "%Y-%m-%d").date()
                    if fv_date <= hoy:
                        estado = 'Vencido'
                    elif fv_date <= limite_crit:
                        estado = 'Alerta Crítica'
                    elif  fv_date <= limite_adv:
                        estado = 'Alerta de Advertencia'
                    elif fv_date <= limite_prev:
                        estado = 'Alerta Preventiva'

                    if estado:
                        alertas.append({
                            'Estado': estado,
                            'Fecha de Vencimiento': fv,
                            'Días restantes': (fv_date - hoy).days,
                            'Nombre': lote['nombre'],
                            'Cantidad del lote': lote['cantidad'],
                            'Código': codigo
                        })
                except:
                    pass


    if alertas:
        df_alertas = pd.DataFrame(alertas).sort_values(by='Días restantes')
        st.dataframe(df_alertas, width='stretch', hide_index=True,
                     column_order=("Código", 'Nombre', 'Cantidad del lote', 'Fecha de Vencimiento', 'Días restantes', 'Estado'))
    else:
        st.write("No hay productos para los rangos seleccionados.")
